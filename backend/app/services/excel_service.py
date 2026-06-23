from datetime import datetime, timezone
from io import BytesIO
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADERS = ["child_id","admission_file_no","full_name","father_name","grandfather_name","mother_name","gender","date_of_birth","guardian_name","guardian_relationship","guardian_cnic","guardian_mobile","current_address","permanent_address","village_mohallah","union_council","tehsil","district","province","admission_date","reason_for_admission","status"]
REQUIRED_HEADERS = {
    "child_id", "admission_file_no", "full_name", "gender", "date_of_birth",
    "guardian_name", "guardian_relationship", "guardian_mobile", "district",
    "province", "admission_date", "reason_for_admission", "status",
}
SAMPLE_DATA = {
    "child_id": "CCMS-0001", "admission_file_no": "AF-2026-0001",
    "full_name": "Ali Khan", "father_name": "Late Muhammad Khan",
    "grandfather_name": "Abdul Rahman", "mother_name": "Ayesha Bibi",
    "gender": "Male", "date_of_birth": "2015-05-15",
    "guardian_name": "Ahmed Khan", "guardian_relationship": "Uncle",
    "guardian_cnic": "17301-1234567-1", "guardian_mobile": "03001234567",
    "current_address": "Mohallah Sheikh Abad, Mardan",
    "permanent_address": "Village Katlang, Mardan", "village_mohallah": "Katlang",
    "union_council": "Katlang", "tehsil": "Mardan", "district": "Mardan",
    "province": "Khyber Pakhtunkhwa", "admission_date": "2026-06-20",
    "reason_for_admission": "Orphan child requiring care and education", "status": "Active",
}

def _finish(ws, header_row, columns):
    fill=PatternFill("solid", fgColor="1F4E78")
    for c in ws[header_row]: c.font=Font(color="FFFFFF",bold=True); c.fill=fill; c.alignment=Alignment(horizontal="center")
    ws.freeze_panes=f"A{header_row+1}"; ws.auto_filter.ref=f"A{header_row}:{get_column_letter(len(columns))}{ws.max_row}"
    for i,name in enumerate(columns,1):
        width=max([len(str(name))]+[len(str(ws.cell(r,i).value or "")) for r in range(header_row+1,ws.max_row+1)])
        ws.column_dimensions[get_column_letter(i)].width=min(width+2,45)

def build_excel_report(title, rows, username, filters):
    wb=Workbook(); ws=wb.active; ws.title="Report"
    ws.append(["CCMS (Child Care Management System)"]); ws.append([title]); ws.append([f"Generated: {datetime.now(timezone.utc).isoformat()} by {username}"]); ws.append([f"Filters: {filters}"])
    columns=list(rows[0]) if rows else ["No data"]
    ws.append(columns)
    for row in rows: ws.append([row.get(c) for c in columns])
    _finish(ws,5,columns); out=BytesIO(); wb.save(out); out.seek(0); return out

def build_child_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Children Import Template"
    ws.append(HEADERS)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"
    ws.sheet_view.showGridLines = False
    required_fill = PatternFill("solid", fgColor="F4B183")
    optional_fill = PatternFill("solid", fgColor="D9EAF7")
    border = Border(bottom=Side(style="thin", color="7F8C8D"))
    for index, header in enumerate(HEADERS, 1):
        cell = ws.cell(1, index)
        cell.font = Font(bold=True, color="1F1F1F")
        cell.fill = required_fill if header in REQUIRED_HEADERS else optional_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        if header in REQUIRED_HEADERS:
            cell.comment = Comment("Required field. Do not leave blank.", "CCMS")
        ws.column_dimensions[get_column_letter(index)].width = min(max(len(header) + 4, 16), 34)
    ws.row_dimensions[1].height = 34
    for row in range(2, 5002):
        ws.cell(row, HEADERS.index("date_of_birth") + 1).number_format = "yyyy-mm-dd"
        ws.cell(row, HEADERS.index("admission_date") + 1).number_format = "yyyy-mm-dd"
    gender_validation = DataValidation(type="list", formula1='"Male,Female,Other"')
    status_validation = DataValidation(type="list", formula1='"Active,Inactive,Discharged,Transferred"')
    ws.add_data_validation(gender_validation); ws.add_data_validation(status_validation)
    gender_validation.add(f"G2:G5001"); status_validation.add(f"V2:V5001")

    instructions = wb.create_sheet("Instructions")
    instructions.append(["CCMS Children Import Instructions"])
    instructions["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    instructions["A1"].alignment = Alignment(vertical="center")
    instructions.row_dimensions[1].height = 28
    instruction_lines = [
        "Do not change column names.", "Enter one child per row.",
        "Use date format YYYY-MM-DD.", "Do not leave required fields blank.",
        "Download Google Sheets as Excel or CSV before upload.", "Use a unique child_id.",
        "Use a unique admission_file_no.", "Preview the import before commit.",
        "Correct all validation errors before committing.",
        "Orange headers are required; blue headers are optional. Maximum 5,000 data rows.",
    ]
    for number, line in enumerate(instruction_lines, 1):
        instructions.append([number, line])
    instructions.column_dimensions["A"].width = 8
    instructions.column_dimensions["B"].width = 85
    for row in instructions.iter_rows(min_row=2, max_col=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")

    allowed = wb.create_sheet("Allowed Values")
    allowed.append(["Field", "Allowed Value"])
    for value in ("Male", "Female", "Other"):
        allowed.append(["gender", value])
    for value in ("Active", "Inactive", "Discharged", "Transferred"):
        allowed.append(["status", value])
    _finish(allowed, 1, ["Field", "Allowed Value"])
    allowed.column_dimensions["A"].width = 20
    allowed.column_dimensions["B"].width = 28

    sample = wb.create_sheet("Sample Data")
    sample.append(HEADERS)
    sample.append([SAMPLE_DATA[header] for header in HEADERS])
    _finish(sample, 1, HEADERS)
    sample.sheet_view.showGridLines = False
    sample.row_dimensions[2].height = 42
    for cell in sample[2]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
