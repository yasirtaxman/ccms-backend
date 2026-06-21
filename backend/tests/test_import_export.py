from io import BytesIO
from openpyxl import Workbook, load_workbook
from app.models.audit_log import AuditLog
from app.services.excel_service import HEADERS, REQUIRED_HEADERS, SAMPLE_DATA
from tests.test_dashboard_reports import auth, child


def upload(rows):
    wb=Workbook(); ws=wb.active; ws.append(HEADERS)
    for row in rows: ws.append(row)
    out=BytesIO(); wb.save(out); return ("children.xlsx",out.getvalue(),"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


VALID=["CH-101","ADM-101","Test Child","Father","Grandfather","Mother","Male","2015-01-01","Guardian","Uncle","00000-0000000-0","03000000000","Current","Permanent","Village","UC","Tehsil","District","Province","2026-01-01","Care","Active"]


def test_template_preview_commit_and_audit(client, db_session):
    _, headers=auth(db_session)
    template=client.get("/imports/templates/children.xlsx",headers=headers)
    assert template.status_code==200
    workbook = load_workbook(BytesIO(template.content))
    assert workbook.sheetnames == ["Children Import Template", "Instructions", "Allowed Values", "Sample Data"]
    template_sheet = workbook["Children Import Template"]
    assert [cell.value for cell in template_sheet[1]] == HEADERS
    assert template_sheet.freeze_panes == "A2"
    for header in REQUIRED_HEADERS:
        cell = template_sheet.cell(1, HEADERS.index(header) + 1)
        assert cell.comment is not None
        assert cell.fill.fgColor.rgb.endswith("F4B183")
    instructions = " ".join(str(cell.value or "") for row in workbook["Instructions"] for cell in row)
    for phrase in ("Do not change column names", "YYYY-MM-DD", "Google Sheets", "Preview", "validation errors"):
        assert phrase in instructions
    allowed = {(row[0].value, row[1].value) for row in workbook["Allowed Values"].iter_rows(min_row=2)}
    assert {("gender", "Male"), ("gender", "Female"), ("gender", "Other")} <= allowed
    assert {("status", "Active"), ("status", "Inactive"), ("status", "Discharged"), ("status", "Transferred")} <= allowed
    sample = workbook["Sample Data"]
    assert [cell.value for cell in sample[2]] == [SAMPLE_DATA[header] for header in HEADERS]
    preview=client.post("/imports/children/preview",headers=headers,files={"file":upload([VALID])})
    assert preview.status_code==200, preview.text
    assert preview.json()["valid_rows"]==1
    committed=client.post("/imports/children/commit",headers=headers,files={"file":upload([VALID])})
    assert committed.status_code==200, committed.text
    assert committed.json()["imported_count"]==1
    assert db_session.query(AuditLog).filter(AuditLog.action.in_(["IMPORT_PREVIEW","IMPORT_COMMIT"])).count()==2


def test_import_validation_duplicates_and_rbac(client, db_session):
    _, data_headers=auth(db_session,"operator","Data Entry Operator")
    duplicate=client.post("/imports/children/preview",headers=data_headers,files={"file":upload([VALID,VALID])})
    assert duplicate.status_code==200 and duplicate.json()["duplicate_rows"]==1
    assert client.post("/imports/children/commit",headers=data_headers,files={"file":upload([VALID])}).status_code==403
    _, viewer_headers=auth(db_session,"viewer2","Viewer")
    assert client.get("/imports/templates/children.xlsx",headers=viewer_headers).status_code==403


def test_optional_fields_may_be_blank_but_required_fields_may_not(client, db_session):
    _, headers = auth(db_session, "optional-fields", "Manager")
    optional_blank = [value if header in REQUIRED_HEADERS else "" for header, value in zip(HEADERS, VALID)]
    valid = client.post("/imports/children/preview", headers=headers, files={"file": upload([optional_blank])})
    assert valid.status_code == 200
    assert valid.json()["valid_rows"] == 1
    missing_required = optional_blank.copy()
    missing_required[HEADERS.index("guardian_mobile")] = ""
    invalid = client.post("/imports/children/preview", headers=headers, files={"file": upload([missing_required])})
    assert invalid.status_code == 200
    assert any(error["field"] == "guardian_mobile" for error in invalid.json()["validation_errors"])


def test_exports_are_downloadable_and_audited(client, db_session):
    _, headers=auth(db_session)
    for path,content_type in (("/exports/children.xlsx","spreadsheetml"),("/exports/children.pdf","application/pdf")):
        response=client.get(path,headers=headers)
        assert response.status_code==200
        assert content_type in response.headers["content-type"]
    assert db_session.query(AuditLog).filter(AuditLog.action.in_(["EXPORT_EXCEL","EXPORT_PDF"])).count()==2


def test_full_child_profile_exports_have_professional_structure(client, db_session):
    _, headers = auth(db_session, "profile-admin", "Admin")
    record = child()
    db_session.add(record)
    db_session.commit()
    excel = client.get(f"/exports/full-child-profile/{record.id}.xlsx", headers=headers)
    assert excel.status_code == 200, excel.text
    workbook = load_workbook(BytesIO(excel.content))
    assert workbook.sheetnames == [
        "Child Basic Info", "Guardian & Address", "Documents", "Sponsorship",
        "Accommodation", "Medical Summary", "Education Summary",
        "Case Management", "Daily Attendance",
    ]
    values = [str(cell.value) for sheet in workbook for row in sheet for cell in row if cell.value is not None]
    assert record.full_name in values
    assert not any("Filters:" in value or "{'child_id'" in value for value in values)
    pdf = client.get(f"/exports/full-child-profile/{record.id}.pdf", headers=headers)
    assert pdf.status_code == 200, pdf.text
    assert pdf.content.startswith(b"%PDF")
    assert len(pdf.content) > 3000


def test_full_child_profile_viewer_export_masks_sensitive_data(client, db_session):
    _, headers = auth(db_session, "profile-viewer", "Viewer")
    record = child()
    db_session.add(record)
    db_session.commit()
    response = client.get(f"/exports/full-child-profile/{record.id}.xlsx", headers=headers)
    assert response.status_code == 200, response.text
    workbook = load_workbook(BytesIO(response.content))
    values = {str(cell.value) for sheet in workbook for row in sheet for cell in row if cell.value is not None}
    assert record.guardian_cnic not in values
    assert record.guardian_mobile not in values
    assert record.current_address not in values
    assert record.permanent_address not in values
